"""Reports module for BMI Calculator.

Handles exporting records to CSV, Excel, and PDF formats.
"""

import csv
import os
from datetime import datetime

from database import get_all_records, get_record_by_id
from utils import RECOMMENDATIONS, format_datetime

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "exports")


def _ensure_exports_dir():
    """Create the exports directory if it doesn't exist."""
    os.makedirs(EXPORTS_DIR, exist_ok=True)


# ── CSV Export ───────────────────────────────────────────────────────────────

def export_to_csv(records: list | None = None) -> str:
    """Export records to CSV. Returns the file path."""
    _ensure_exports_dir()
    if records is None:
        records = get_all_records()
    if not records:
        raise ValueError("No records to export.")

    filename = f"bmi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(EXPORTS_DIR, filename)

    headers = ["ID", "Name", "Age", "Gender", "Weight (kg)",
               "Height (cm)", "BMI", "Category", "Date & Time"]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in records:
            writer.writerow([
                r["id"], r["name"], r["age"], r["gender"],
                r["weight"], r["height"], r["bmi"],
                r["category"], r["date_time"],
            ])

    return filepath


# ── Excel Export ─────────────────────────────────────────────────────────────

def export_to_excel(records: list | None = None) -> str:
    """Export records to Excel. Returns the file path."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    _ensure_exports_dir()
    if records is None:
        records = get_all_records()
    if not records:
        raise ValueError("No records to export.")

    filename = f"bmi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BMI Records"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["ID", "Name", "Age", "Gender", "Weight (kg)",
               "Height (cm)", "BMI", "Category", "Date & Time"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(records, 2):
        values = [r["id"], r["name"], r["age"], r["gender"],
                  r["weight"], r["height"], r["bmi"],
                  r["category"], r["date_time"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_len

    wb.save(filepath)
    return filepath


# ── PDF Export ───────────────────────────────────────────────────────────────

def export_to_pdf(record_id: int | None = None) -> str:
    """Export a single record or all records to PDF. Returns the file path."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError(
            "fpdf2 is required for PDF export. "
            "Install it with: pip install fpdf2"
        )

    _ensure_exports_dir()

    if record_id is not None:
        rec = get_record_by_id(record_id)
        if not rec:
            raise ValueError(f"Record {record_id} not found.")
        records = [rec]
    else:
        records = get_all_records()
        if not records:
            raise ValueError("No records to export.")

    filename = f"bmi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(EXPORTS_DIR, filename)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for rec in records:
        pdf.add_page()

        # Title
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_text_color(44, 62, 80)
        pdf.cell(0, 12, "BMI Health Report", ln=True, align="C")
        pdf.ln(4)

        # Separator
        pdf.set_draw_color(52, 152, 219)
        pdf.set_line_width(0.8)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(8)

        # User Details
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(52, 152, 219)
        pdf.cell(0, 8, "User Details", ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(44, 62, 80)

        details = [
            ("Full Name", rec["name"]),
            ("Age", str(rec["age"])),
            ("Gender", rec["gender"]),
            ("Weight", f"{rec['weight']} kg"),
            ("Height", f"{rec['height']} cm"),
            ("Date & Time", rec["date_time"]),
        ]
        for label, value in details:
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(50, 7, f"{label}:", align="L")
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 7, value, ln=True)

        pdf.ln(6)

        # BMI Result
        pdf.set_draw_color(52, 152, 219)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(52, 152, 219)
        pdf.cell(0, 8, "BMI Result", ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 28)
        pdf.cell(0, 14, f"{rec['bmi']:.2f}", ln=True, align="C")

        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 8, rec["category"], ln=True, align="C")
        pdf.ln(4)

        # Recommendation
        pdf.set_draw_color(52, 152, 219)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(52, 152, 219)
        pdf.cell(0, 8, "Recommendation", ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(44, 62, 80)
        recommendation = RECOMMENDATIONS.get(rec["category"], "")
        pdf.multi_cell(0, 6, recommendation)

        # Footer
        pdf.ln(10)
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(127, 140, 141)
        pdf.cell(0, 5, f"Report generated on {format_datetime()}", ln=True,
                 align="C")
        pdf.cell(0, 5, "BMI Health Calculator - For informational purposes only",
                 ln=True, align="C")

    pdf.output(filepath)
    return filepath
